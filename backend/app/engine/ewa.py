"""Parse an SAP EarlyWatch Alert (EWA) export into a partial STAR intake plus an
extraction summary, mirroring the Readiness Check parser contract
(``{"form": {...}, "summary": {...}}``).

EWA is delivered as the classic session HTML report, as XLSX, or as PDF. All
three are reduced to one plain-text blob and fed through the same regex
extractor — PDFs use ``pypdf`` (pure-Python, no system deps).

Heuristic by nature: EWA layout shifts across SAP_BASIS releases, so we de-tag
the HTML to plain text and regex for the stable phrases ("Database Size",
"growth", the performance rating, the average dialog response time) rather than
relying on fixed cell coordinates. EWA contributes two intake fields STAR scores
on — ``db_size_band`` and ``stack_type`` — plus narrative facts and an advisory.
"""
import html as _html
import io
import re
import zipfile
from typing import Optional

# Reuse the Readiness Check size->band mapping so both parsers band DB size identically.
from app.engine.readiness import _band_from_gb

_UNIT_TO_GB = {"mb": 1 / 1024, "gb": 1.0, "tb": 1024.0}


def _num(s: str) -> float:
    """Parse a number that may use US ('1,234.5') or EU ('1.234,5') conventions.

    EWA reports localise to the customer's locale, so DB size shows up as both
    '3675.77' (US) and '3.675,77' (EU) across sections. The last separator is
    treated as the decimal mark; the other separates thousands.
    """
    s = s.strip()
    if not s:
        return 0.0
    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            return float(s.replace(".", "").replace(",", "."))
        return float(s.replace(",", ""))
    if "," in s:
        return float(s.replace(",", "."))
    return float(s)


def _to_gb(value: str, unit: str) -> float:
    """Normalise a "1,234.5 GB"-style figure to GiB."""
    return _num(value) * _UNIT_TO_GB.get(unit.lower(), 1.0)


def _detag(raw: str) -> str:
    """Strip an EWA HTML report down to collapsed, entity-decoded plain text."""
    txt = re.sub(r"(?is)<(script|style).*?</\1>", " ", raw)
    txt = re.sub(r"(?s)<[^>]+>", " ", txt)
    txt = _html.unescape(txt)
    return re.sub(r"[ \t\r\n]+", " ", txt).strip()


def _text_from_xlsx(data: bytes) -> str:
    """Flatten every cell of every sheet into one space-joined string. Reusing
    the same regexes as the HTML path keeps the two formats in lock-step."""
    from openpyxl import load_workbook  # imported lazily; only XLSX needs it

    wb = load_workbook(io.BytesIO(data), read_only=False, data_only=True)
    cells = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            for c in row:
                if c is not None and str(c).strip():
                    cells.append(str(c).strip())
    return " ".join(cells)


def _looks_like_zip(data: bytes) -> bool:
    return data[:2] == b"PK"


def _text_from_pdf(data: bytes) -> str:
    """Extract and collapse all page text from a PDF EWA export."""
    from pypdf import PdfReader  # lazy import; only PDFs need it

    reader = PdfReader(io.BytesIO(data))
    pages = []
    for page in reader.pages:
        try:
            pages.append(page.extract_text() or "")
        except Exception:
            continue
    return re.sub(r"[ \t\r\n]+", " ", " ".join(pages)).strip()


def _extract(text: str) -> dict:
    """Pull the EWA facts out of de-tagged text. Returns a dict of the raw
    figures we found (any subset may be missing on a given report)."""
    out: dict = {}

    # --- production database size ---
    m = re.search(
        r"(?i)(?:size of the database|database size|db size|total\s+db\s+size)"
        r"[^\d]{0,40}([\d.,]+)\s*(tb|gb|mb)",
        text,
    )
    if m:
        out["db_gb"] = _to_gb(m.group(1), m.group(2))

    # --- growth: prefer the explicit "database growth of X% per year" DVM phrase,
    #     else fall back to a looser "growth ... X%" match. Cap at 200% to reject
    #     unrelated percentages the looser pattern may sweep up. ---
    m = re.search(r"(?i)database growth of\s*([\d.,]+)\s*%\s*per\s*year", text)
    if not m:
        m = re.search(r"(?i)(?:growth|grow[ns]?\s+by|increase[sd]?\s+by)[^.%]{0,40}?([\d.,]+)\s*%", text)
    if m:
        g = _num(m.group(1))
        if 0 <= g < 200:
            out["growth_pct"] = g

    # --- overall performance rating (EWA flags GOOD / FAIR / POOR or a traffic light) ---
    # Allow a few words (e.g. "Performance Overview") between the cue and the rating.
    m = re.search(r"(?i)performance\b[\s\S]{0,40}?\b(very good|good|fair|poor|bad|critical)\b", text)
    if m:
        out["perf_rating"] = m.group(1).title()
    else:
        m = re.search(r"(?i)performance\b[\s\S]{0,40}?\b(green|yellow|amber|red)\b", text)
        if m:
            out["perf_rating"] = {"green": "Good", "yellow": "Fair", "amber": "Fair",
                                  "red": "Poor"}[m.group(1).lower()]

    # --- average dialog response time (ms) ---
    # EWA writes this as "Avg. Response Time in Dialog Task NNN ms" — the qualifier
    # follows the metric name rather than preceding it, so allow either order.
    m = re.search(
        r"(?i)(?:avg\.?|average)\s+(?:dialog\s+)?response\s+time"
        r"(?:\s+in\s+dialog(?:\s+task)?)?[^\d]{0,30}([\d.,]+)\s*ms",
        text,
    )
    if m:
        out["dialog_ms"] = int(round(_num(m.group(1))))

    # --- dual-stack (ABAP+Java) detection ---
    if re.search(r"(?i)dual[\s-]?stack|abap\s*\+\s*java|abap\s+and\s+java", text):
        out["dual_stack"] = True

    # --- top growth / largest tables ---
    top = _extract_top_tables(text)
    if top:
        out["top_tables"] = top

    return out


# Tokens that look like SAP table names but are SQL/EWA boilerplate.
_TOP_STOP = {"SAP", "GB", "TB", "MB", "DB", "THE", "AND", "FOR", "SIZE", "ABAP",
             "HANA", "TABLE", "INDEX", "CLUSTER", "LOBSEGMENT", "SYSTEM", "TOTAL", "DATE"}


def _extract_top_tables(text: str) -> list:
    """Find a "Top N Tables" / "Top N Segments" block and return ``[(name, gb), ...]``.

    EWA reports use several layouts depending on DB platform (Oracle/HANA/MaxDB)
    and SAP_BASIS version. We try them in priority order, stopping at the first
    block that yields rows:

    A. ``Top 10 Tables`` — ``NAME size table_size index_size lob_size pct cum_pct``
       (six numbers after the name; size is the first).
    B. ``Top N Segments based on monthly growth rate`` — most relevant signal —
       layout ``NAME TYPE TABLESPACE size extents``.
    C. ``Top N Segments based on size`` / ``... based on extents`` — same layout as B.
    D. Legacy ``largest tables`` / ``top tables`` heading with explicit ``NN GB``.
    """
    # A. "Top 10 Tables" — 6 numbers after the table name.
    h = re.search(r"(?i)Top\s+10\s+Tables\b", text)
    if h:
        rows = _collect_rows(
            text[h.start(): h.start() + 2500],
            r"\b([A-Z][A-Z0-9_/~]{2,29})\b\s+([\d.,]+)"
            r"\s+[\d.,]+\s+[\d.,]+\s+[\d.,]+\s+[\d.,]+\s+[\d.,]+",
        )
        if rows:
            return rows

    # B. "Top N Segments based on monthly growth rate" — growth signal beats size.
    for heading in (
        r"Top\s+\d+\s+Segments\s+based\s+on\s+monthly\s+growth\s+rate",
        r"Top\s+\d+\s+Segments\s+based\s+on\s+size",
        r"Top\s+\d+\s+Segments\b",
    ):
        h = re.search(r"(?i)" + heading, text)
        if not h:
            continue
        rows = _collect_rows(
            text[h.start(): h.start() + 2500],
            r"\b([A-Z][A-Z0-9_/~$]{2,29})\b\s+(?:TABLE|INDEX|CLUSTER|LOBSEGMENT)"
            r"\s+\S+\s+([\d.,]+)\s+\d+",
        )
        if rows:
            return rows

    # D. Legacy "largest tables" / "top tables" with explicit "GB" per row.
    h = re.search(r"(?i)(largest tables|top\s+(?:size\s+consumers|growth\s+tables?|tables?))", text)
    if h:
        rows = _collect_rows(
            text[h.start(): h.start() + 1500],
            r"\b([A-Z][A-Z0-9_/]{2,29})\b[^A-Za-z\n]{0,20}?([\d.,]+)\s*GB",
        )
        if rows:
            return rows

    return []


def _collect_rows(region: str, pattern: str) -> list:
    """Apply ``pattern`` over ``region`` and reduce duplicate table names to the
    largest GB figure. Returns the top 6 sorted by size desc."""
    seen: dict = {}
    for m in re.finditer(pattern, region):
        name, gb = m.group(1), _num(m.group(2))
        if name in _TOP_STOP or gb <= 0:
            continue
        seen[name] = max(seen.get(name, 0.0), gb)
    return sorted(seen.items(), key=lambda kv: kv[1], reverse=True)[:6]


def parse_ewa(data: bytes, filename: str = "") -> dict:
    """Return ``{'form': {...partial intake...}, 'summary': {...}, 'insights': {...}}``.

    Raises ``ValueError`` for unsupported/unparseable input so the API layer can
    surface a 422 with a clear message.
    """
    name = (filename or "").lower()
    if name.endswith(".pdf") or data[:4] == b"%PDF":
        text = _text_from_pdf(data)
    elif name.endswith((".xlsx", ".xlsm")) or _looks_like_zip(data):
        text = _text_from_xlsx(data)
    else:
        raw = data.decode("utf-8", "ignore") if b"\x00" not in data[:64] else data.decode("utf-16", "ignore")
        text = _detag(raw)

    facts_in = _extract(text)
    if not facts_in:
        raise ValueError("No EarlyWatch metrics found — is this an EWA HTML/XLSX/PDF export?")

    form: dict = {}
    extracted: dict = {}
    facts: list = []
    advisory: Optional[str] = None
    review = ["DB growth driver tables vs archiving candidates",
              "Performance findings beyond the headline rating",
              "Business inputs: driver, go-live, budget, risk, sovereignty, basis capability"]

    db_gb = facts_in.get("db_gb")
    growth = facts_in.get("growth_pct")
    if db_gb:
        form["db_size_band"] = _band_from_gb(db_gb)
        extracted["db_size_gb"] = int(round(db_gb))
        g = f", growth ≈ {growth:g}%/yr" if growth is not None else ""
        facts.append(f"Production DB ≈ {int(db_gb)} GB{g}")

    if facts_in.get("dual_stack"):
        form["stack_type"] = "dual_stack"
        facts.append("Dual-stack (ABAP+Java) detected — split before any S/4HANA conversion")
    else:
        # EWA describes a single ABAP system unless it flags Java — record the inference.
        form["stack_type"] = "single_stack"

    rating = facts_in.get("perf_rating")
    dialog = facts_in.get("dialog_ms")
    if rating or dialog:
        bits = []
        if rating:
            bits.append(f"performance {rating.upper()}")
        if dialog:
            bits.append(f"avg dialog {dialog} ms")
        facts.append("System health: " + ", ".join(bits))

    top = facts_in.get("top_tables") or []
    if top:
        facts.append("Top growth tables: " + " · ".join(f"{n} ({int(gb)} GB)" for n, gb in top))

    # Advisory: meaningful growth means archive/DVM before a conversion migration.
    if growth is not None and growth >= 5:
        advisory = (f"EarlyWatch shows ≈ {growth:g}%/yr DB growth"
                    + (" concentrated in " + top[0][0] if top else "")
                    + " — run data volume management / archiving before conversion to cut "
                    "migration runtime and downtime.")
        facts.append("Advisory: archive / DVM before conversion to shrink migration runtime")

    insights = {
        "kind": "ewa",
        "db": {"sizeGb": int(db_gb) if db_gb else None, "growthPct": growth},
        "perf": {"rating": rating, "dialogMs": dialog},
        "topTables": [[n, int(gb)] for n, gb in top],
    }

    return {"form": form, "extracted": extracted,
            "summary": {"facts": facts, "review": review, "advisory": advisory},
            "insights": insights}
