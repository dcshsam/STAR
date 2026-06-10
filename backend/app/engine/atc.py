"""Parse an SAP ATC / Custom Code Check export into a partial STAR intake.

Two formats are supported, auto-detected from the header row:

RC Custom Code format (CustomCode.xlsx from the SAP Readiness Check export):
  Columns: Custom Code Topic | Status | Type of Remediation | Quick Fix Support |
           In Scope | Out of Scope | Exemption State | Usage Information |
           Priority | Number of Custom Objects | Application Component | SAP Note

Pure ATC / SCI format (export from transaction SCI / ABAP Test Cockpit):
  Columns: Object | Object Type | Check | Severity | In Scope | Used | ...

Both return the same contract: {"form": {...}, "summary": {...}, "insights": {...}}
"""
import io
import re
import zipfile
from collections import Counter
from typing import Optional

from openpyxl import load_workbook

# ── Application Component prefix → STAR PRESET_MODULES ───────────────────────
_COMP_TO_MODULE: list[tuple[str, str]] = [
    ("fi-", "FI"), ("fi/", "FI"),
    ("co-", "CO"), ("co/", "CO"),
    ("fscm", "FSCM"),
    ("mm-", "MM"), ("mm/", "MM"),
    ("srm", "SRM"), ("mm-srv", "SRM"),
    ("sd-", "SD"), ("sd/", "SD"),
    ("pp-", "PP"), ("pp/", "PP"),
    ("qm-", "QM"),
    ("pm-", "PM / EAM"), ("cs-", "PM / EAM"),
    ("ps-", "PS"),
    ("hr-", "HCM"), ("pa-", "HCM"), ("py-", "HCM"), ("pt-", "HCM"), ("hcm", "HCM"),
    ("le-wm", "WM"), ("wm-", "WM"),
    ("ewm", "EWM"),
    ("tm-", "TM"),
    ("grc", "GRC"),
    ("re-", "RE"),
]


def _comp_to_modules(comp: str) -> list[str]:
    cl = comp.lower()
    return [mod for prefix, mod in _COMP_TO_MODULE if cl.startswith(prefix)]


def _band(total: int) -> str:
    if total < 500:
        return "lt_500"
    if total < 2000:
        return "500_2k"
    if total <= 10000:
        return "2k_10k"
    return "gt_10k"


def _rows(xlsx_bytes: bytes) -> list:
    wb = load_workbook(io.BytesIO(xlsx_bytes), read_only=False, data_only=True)
    ws = wb.worksheets[0]
    return [r for r in ws.iter_rows(values_only=True)
            if any(c is not None and str(c).strip() for c in r)]


def _col(headers: list, *keywords: str) -> Optional[int]:
    for kw in keywords:
        for i, h in enumerate(headers):
            if kw in str(h).lower():
                return i
    return None


def _to_int(s) -> int:
    if s is None:
        return 0
    try:
        return int(float(str(s).replace(",", "").strip()))
    except (ValueError, TypeError):
        return 0


def _parse_priority_str(s: str) -> tuple[int, int, int]:
    """Parse 'Errors-63, Warnings-759, Information-1896' → (63, 759, 1896)."""
    def _n(label: str) -> int:
        m = re.search(label + r"[-:]\s*(\d+)", s, re.I)
        return int(m.group(1)) if m else 0
    return _n("errors"), _n("warnings"), _n("information")


# ── RC Custom Code format ─────────────────────────────────────────────────────

def _is_rc_cc_format(headers: list) -> bool:
    joined = " | ".join(headers).lower()
    return "custom code topic" in joined or (
        "type of remediation" in joined and "application component" in joined
    )


def _parse_rc_cc(data_rows: list, headers: list) -> dict:
    """Parse the RC CustomCode.xlsx topic-level format."""
    topic_col  = _col(headers, "custom code topic", "topic", "title")
    status_col = _col(headers, "status")
    remed_col  = _col(headers, "type of remediation", "remediation")
    scope_col  = _col(headers, "in scope")
    obj_col    = _col(headers, "number of custom objects", "num", "objects")
    prio_col   = _col(headers, "priority")
    comp_col   = _col(headers, "application component", "component")
    usage_col  = _col(headers, "usage information", "usage")
    exempt_col = _col(headers, "exemption state", "exemption")
    qf_col     = _col(headers, "quick fix support", "quick fix")

    def cell(row, col):
        if col is not None and col < len(row) and row[col] is not None:
            return str(row[col]).strip()
        return ""

    total_objects = 0          # sum of "Number of Custom Objects"
    total_in_scope = 0         # sum of "In Scope"
    total_errors = 0
    total_warnings = 0
    total_unresolved = 0
    functional_redesign = 0
    unavailable = 0
    qf_objects = 0             # sum of N from "Quick Fix Available - N"
    qf_topics = 0              # # of topics with any quick-fix support
    qf_not_avail_topics = 0    # # of topics marked "Quick Fix Not Available"
    modules: list[str] = []
    remed_counter: Counter = Counter()
    status_counter: Counter = Counter()
    comp_counter: Counter = Counter()

    for row in data_rows:
        n_obj  = _to_int(cell(row, obj_col))
        n_scope = _to_int(cell(row, scope_col))
        total_objects += n_obj if n_obj else 1
        total_in_scope += n_scope

        # Priority string: "Errors-63, Warnings-759, Information-1896"
        prio_str = cell(row, prio_col)
        if prio_str:
            e, w, _ = _parse_priority_str(prio_str)
            total_errors += e
            total_warnings += w

        # Remediation type
        remed = cell(row, remed_col)
        if remed:
            remed_counter[remed] += 1
            if "functional" in remed.lower() or "redesign" in remed.lower():
                functional_redesign += 1

        # Status
        status = cell(row, status_col)
        if status:
            sl = status.lower()
            status_counter[status] += 1
            if "unavailable" in sl:
                unavailable += 1

        # Unresolved count from exemption state
        exempt_str = cell(row, exempt_col)
        m = re.search(r"unresolved[-:]\s*(\d+)", exempt_str, re.I)
        if m:
            total_unresolved += int(m.group(1))

        # Quick Fix Support: "Quick Fix Available - 2468" or "Quick Fix Not Available"
        qf_str = cell(row, qf_col).lower()
        if qf_str:
            qm = re.search(r"quick fix available[^\d]*(\d+)", qf_str)
            if qm:
                qf_objects += int(qm.group(1))
                qf_topics += 1
            elif "not available" in qf_str:
                qf_not_avail_topics += 1

        # Application Component → modules
        comp = cell(row, comp_col)
        if comp:
            comp_counter[comp.split("/")[0].strip()] += 1
            for mod in _comp_to_modules(comp):
                if mod not in modules:
                    modules.append(mod)

    total_topics = len(data_rows)

    # Derive form fields
    form: dict = {
        "custom_objects_band": _band(total_objects),
        "modifications_to_standard": "true",
    }
    if modules:
        form["modules_implemented"] = modules

    # Overall customization from error volume
    if total_errors >= 500 or total_objects >= 5000:
        form["overall_customization"] = "High"
    elif total_errors >= 100 or total_objects >= 1000:
        form["overall_customization"] = "Med"
    else:
        form["overall_customization"] = "Low"

    # Process reengineering from functional redesign items + unavailable
    redesign_pct = functional_redesign / total_topics if total_topics else 0
    if unavailable >= 5 or redesign_pct >= 0.3:
        form["process_reengineering_appetite"] = "redesign_to_standard"
    elif unavailable >= 2 or redesign_pct >= 0.1:
        form["process_reengineering_appetite"] = "selective"

    # Quick-fix coverage discounts manual remediation effort.
    qf_coverage_pct = round(qf_objects / total_in_scope * 100) if total_in_scope else 0
    base_effort = total_errors * 2.5 + total_warnings * 0.5
    qf_discount = min(qf_coverage_pct / 100 * 0.7, 0.7)   # quick-fix saves up to 70% of the manual effort
    qf_automation_pd = qf_objects * 0.05                  # apply + validate the auto-fix
    effort_pd = round(base_effort * (1 - qf_discount) + qf_automation_pd, 0)
    effort_str = f"{int(effort_pd * 0.8)}–{int(effort_pd * 1.2)} PD" if effort_pd else "< 5 PD"

    facts = [
        f"{total_topics} custom-code topics · {total_objects:,} custom objects · "
        f"{total_in_scope:,} in scope",
        f"{total_errors:,} errors + {total_warnings:,} warnings across all topics",
    ]
    if total_unresolved:
        facts.append(f"{total_unresolved:,} objects still unresolved (not in baseline or exempted)")
    if qf_objects or qf_not_avail_topics:
        manual_objects = max(total_in_scope - qf_objects, 0)
        facts.append(
            f"Quick Fix: {qf_objects:,} objects auto-fixable across {qf_topics} topic(s) "
            f"({qf_coverage_pct}% coverage) · {manual_objects:,} objects need manual remediation "
            f"({qf_not_avail_topics} topic(s) without Quick Fix support)"
        )
    if functional_redesign:
        facts.append(
            f"{functional_redesign} topic(s) require Functional Redesign "
            f"({unavailable} involve Functionality Unavailable)"
        )
    if modules:
        facts.append(f"Affected SAP modules: {', '.join(modules)}")
    facts.append(f"Remediation estimate ≈ {effort_str}")

    advisory: Optional[str] = None
    if unavailable >= 3:
        advisory = (
            f"RC Custom Code check flags {unavailable} topic(s) as "
            "'Functionality Unavailable' — these require Functional Redesign "
            "before conversion and will affect process scope."
        )
    elif total_errors >= 200:
        advisory = (
            f"{total_errors:,} error-level findings detected — schedule "
            "custom-code remediation before the conversion freeze."
        )

    by_remed   = remed_counter.most_common()
    by_status  = status_counter.most_common()
    by_comp    = comp_counter.most_common(10)

    insights = {
        "kind": "atc",
        "format": "rc_cc",
        "totals": {
            "topics": total_topics,
            "objects": total_objects,
            "inScope": total_in_scope,
            "errors": total_errors,
            "warnings": total_warnings,
            "unresolved": total_unresolved,
            "functionalRedesign": functional_redesign,
            "unavailable": unavailable,
            "usedPct": None,
            "effort": effort_str,
        },
        "quickFix": {
            "objectsAvailable":  qf_objects,
            "objectsManual":     max(total_in_scope - qf_objects, 0),
            "topicsAvailable":   qf_topics,
            "topicsNotAvail":    qf_not_avail_topics,
            "coveragePct":       qf_coverage_pct,
        },
        "byCategory": by_remed,
        "byStatus": by_status,
        "byComponent": by_comp,
    }
    review = [
        "Functional Redesign topics — confirm scope and process owners",
        "Unresolved objects — confirm baseline / exemption plan before conversion freeze",
        "Effort estimate breakdown by module and team",
        "Business inputs: driver, go-live, budget, risk, sovereignty, basis capability",
    ]
    extracted = {
        "custom_objects": total_in_scope or total_objects,
        "quick_fix_objects": qf_objects,
        "quick_fix_pct": qf_coverage_pct,
        "manual_objects": max(total_in_scope - qf_objects, 0),
    }
    return {"form": form, "summary": {"facts": facts, "review": review,
                                       "advisory": advisory}, "insights": insights,
            "extracted": extracted}


# ── Pure ATC / SCI format ─────────────────────────────────────────────────────

_SEV_KWS   = ["severity", "prio", "priority", "message type", "check priority"]
_IN_SCOPE_KWS = ["in scope", "inscope", "scope", "relevant"]
_ACTIVE_KWS = ["used", "active", "upl", "in use", "actively used"]
_CAT_KWS   = ["check name", "check id", "check", "category", "finding"]

_SEV_MAP = {
    "e": "error", "error": "error", "1": "error",
    "w": "warning", "warning": "warning", "2": "warning",
    "i": "info", "info": "info", "3": "info",
}


def _parse_sci(data_rows: list, headers: list) -> dict:
    """Parse the classic SCI / ABAP Test Cockpit one-finding-per-row format."""
    sev_col    = _col(headers, *_SEV_KWS)
    scope_col  = _col(headers, *_IN_SCOPE_KWS)
    active_col = _col(headers, *_ACTIVE_KWS)
    cat_col    = _col(headers, *_CAT_KWS)

    errors = warnings = infos = 0
    in_scope_total = active_total = total_objects = 0
    cat_counter: Counter = Counter()

    for row in data_rows:
        total_objects += 1
        sev_raw = str(row[sev_col]).strip().lower() if sev_col is not None and sev_col < len(row) else ""
        mapped = _SEV_MAP.get(sev_raw[:1], _SEV_MAP.get(sev_raw, "info"))
        if mapped == "error":
            errors += 1
        elif mapped == "warning":
            warnings += 1
        else:
            infos += 1

        if cat_col is not None and cat_col < len(row) and row[cat_col]:
            raw_cat = str(row[cat_col]).strip()[:60]
            if raw_cat and raw_cat.lower() not in ("none", ""):
                cat_counter[raw_cat] += 1

        if scope_col is not None and scope_col < len(row):
            v = row[scope_col]
            try:
                in_scope_total += int(float(str(v).replace(",", "").strip()))
            except (TypeError, ValueError):
                if str(v).strip().lower() in ("x", "yes", "true", "1"):
                    in_scope_total += 1

        if active_col is not None and active_col < len(row):
            v = row[active_col]
            try:
                active_total += int(float(str(v).replace(",", "").strip()))
            except (TypeError, ValueError):
                if str(v).strip().lower() in ("x", "yes", "true", "1"):
                    active_total += 1

    pct_active = round((active_total / total_objects * 100) if total_objects else 50)
    total_findings = errors + warnings
    effort_pd = round(errors * 2.5 + warnings * 0.5, 0)
    effort_str = f"{int(effort_pd * 0.8)}–{int(effort_pd * 1.2)} PD" if effort_pd else "< 5 PD"

    form: dict = {
        "custom_objects_band": _band(in_scope_total if in_scope_total else total_objects),
        "overall_customization": ("High" if errors >= 40 or total_findings >= 200
                                  else "Med" if errors >= 12 or total_findings >= 50
                                  else "Low"),
        "modifications_to_standard": "true",
    }
    if active_col is not None:
        form["pct_active_estimate"] = min(max(pct_active, 0), 100)

    advisory: Optional[str] = None
    if errors >= 40:
        advisory = (f"ATC found {errors:,} mandatory (error-level) findings — schedule "
                    "custom-code remediation before the conversion freeze.")

    facts = [
        f"{total_objects:,} custom objects analysed"
        + (f" · {in_scope_total:,} in scope" if in_scope_total else ""),
        f"{errors:,} error-level + {warnings:,} warning findings",
    ]
    if active_col is not None:
        facts.append(f"≈ {pct_active}% of custom code actively used (UPL)")
    facts.append(f"Remediation estimate ≈ {effort_str}")

    insights = {
        "kind": "atc",
        "format": "sci",
        "totals": {
            "objects": total_objects,
            "inScope": in_scope_total,
            "errors": errors,
            "warnings": warnings,
            "usedPct": pct_active if active_col is not None else None,
            "effort": effort_str,
        },
        "byCategory": cat_counter.most_common(10),
        "byStatus": [],
        "byComponent": [],
    }
    extracted = {"custom_objects": in_scope_total or total_objects}
    return {"form": form, "summary": {"facts": facts, "review": [
        "Top error-finding categories and owner teams",
        "Effort estimate breakdown by module",
        "Business inputs: driver, go-live, budget, risk, sovereignty, basis capability",
    ], "advisory": advisory}, "insights": insights, "extracted": extracted}


# ── Public entry point ────────────────────────────────────────────────────────

def parse_atc(data: bytes, filename: str = "") -> dict:
    """Return ``{'form': {...}, 'summary': {...}, 'insights': {...}}``.

    Forgiving: if the user uploads a full SAP Readiness Check ZIP here, detect it
    (presence of InterfaceImpactAnalysis*.xlsx) and delegate to the RC parser so
    they get the full picture (interfaces + Fiori + custom code) even though they
    clicked the ATC tile.
    """
    raw = data
    if data[:2] == b"PK":
        try:
            z = zipfile.ZipFile(io.BytesIO(data))
            all_names = z.namelist()
            xlsx_names = [n for n in all_names
                          if n.lower().endswith((".xlsx", ".xlsm"))
                          and not n.split("/")[-1].startswith("~$")]

            # Full RC export? → use the RC parser instead.
            if any("interfaceimpactanalysis" in n.lower() for n in xlsx_names):
                from app.engine.readiness import parse_zip as _rc_parse
                return _rc_parse(data)

            # Otherwise prefer CustomCode.xlsx over the alphabetically first file.
            if xlsx_names:
                cc = next((n for n in xlsx_names if n.lower().endswith("customcode.xlsx")), None)
                raw = z.read(cc or xlsx_names[0])
        except zipfile.BadZipFile:
            pass

    rows = _rows(raw)
    if not rows:
        raise ValueError("ATC export appears empty")

    hdr_idx = 0
    headers: list = []
    for i, row in enumerate(rows[:8]):
        vals = [str(c).strip() for c in row if c is not None]
        if len(vals) >= 3 and not any(v.lstrip("-").replace(".", "").isdigit() for v in vals[:4]):
            headers = [str(c).lower().strip() if c else "" for c in row]
            hdr_idx = i
            break

    data_rows = rows[hdr_idx + 1:]

    if _is_rc_cc_format(headers):
        return _parse_rc_cc(data_rows, headers)
    return _parse_sci(data_rows, headers)
