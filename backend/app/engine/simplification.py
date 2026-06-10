"""Parse an SAP Simplification Item Check export (XLSX or ZIP of XLSX) into a
partial STAR intake plus an extraction summary.

Column names from real SAP Readiness Check exports:
  Title | Effort Ranking | Category | Relevance | LoB/Technology | Business Area |
  Consistency Status | Manual Status | Business Impact Note | Relevance Summary |
  ID | GUID | Comments

Priority is derived from Relevance + Category:
  - "Relevance to Be Checked"          → check (low)
  - "Relevant" + "unavailable"         → high (mandatory)
  - "Relevant" + "change/deprecated"   → medium
"""
import io
import re
import zipfile
from collections import Counter
from typing import Optional

from openpyxl import load_workbook


def _rows(xlsx_bytes: bytes) -> list:
    wb = load_workbook(io.BytesIO(xlsx_bytes), read_only=False, data_only=True)
    ws = wb.worksheets[0]
    return [r for r in ws.iter_rows(values_only=True)
            if any(c is not None and str(c).strip() for c in r)]


def _find_col(headers: list, *candidates: str) -> Optional[int]:
    """Return first column index whose header matches any candidate (substring, lowercase)."""
    for cand in candidates:
        for i, h in enumerate(headers):
            if cand in str(h).lower():
                return i
    return None


# LoB label (lowercase substring) → SAP module codes
# Maps LoB label substrings → PRESET_MODULES names (must match frontend exactly)
_LOB_TO_MODULES: list[tuple[str, list[str]]] = [
    ("finance",                    ["FI", "CO", "FSCM"]),
    ("human resources",            ["HCM"]),
    ("sourcing and procurement",   ["MM", "SRM"]),
    ("sales",                      ["SD"]),
    ("supply chain",               ["PP", "WM", "EWM", "TM"]),
    ("manufacturing",              ["PP"]),
    ("quality",                    ["QM"]),
    ("plant maintenance",          ["PM / EAM"]),
    ("project",                    ["PS"]),
    ("customer service",           ["SD"]),
    ("real estate",                ["RE"]),
    ("governance",                 ["GRC"]),
]


def _lob_to_modules(lob_counter: Counter) -> list[str]:
    mods: list[str] = []
    for lob, _ in lob_counter.items():
        lob_l = lob.lower()
        for key, codes in _LOB_TO_MODULES:
            if key in lob_l:
                for c in codes:
                    if c not in mods:
                        mods.append(c)
    return mods


def _priority(category: str, relevance: str) -> int:
    """0=high, 1=medium, 2=check. Based on actual SAP RC column semantics."""
    r = relevance.lower()
    c = category.lower()
    if "to be checked" in r or "cannot" in r:
        return 2
    if "relevant" in r:
        if "unavailable" in c:
            return 0
        if "change" in c or "deprecated" in c:
            return 1
    return 2


def parse_simplification(data: bytes, filename: str = "") -> dict:
    """Return ``{'form': {...}, 'summary': {...}, 'insights': {...}}``."""
    raw = data
    if data[:2] == b"PK":
        try:
            z = zipfile.ZipFile(io.BytesIO(data))
            xlsx_names = [n for n in z.namelist() if n.lower().endswith((".xlsx", ".xlsm"))]
            if xlsx_names:
                raw = z.read(xlsx_names[0])
        except zipfile.BadZipFile:
            pass

    rows = _rows(raw)
    if not rows:
        raise ValueError("Simplification Item export appears empty")

    # Detect header row (first row with >= 2 non-numeric cells in first 3 cols)
    hdr_idx = 0
    headers: list = []
    for i, row in enumerate(rows[:8]):
        vals = [str(c).strip() for c in row if c is not None]
        if len(vals) >= 2 and not any(v.lstrip("-").replace(".", "").isdigit() for v in vals[:3]):
            headers = [str(c).lower().strip() if c else "" for c in row]
            hdr_idx = i
            break

    data_rows = rows[hdr_idx + 1:]

    title_col    = _find_col(headers, "title", "description", "simplification item", "name")
    category_col = _find_col(headers, "category")
    relevance_col= _find_col(headers, "relevance")
    lob_col      = _find_col(headers, "lob", "technology", "line of business")
    area_col     = _find_col(headers, "business area", "business unit")
    status_col   = _find_col(headers, "consistency status", "consistency", "check status")
    summary_col  = _find_col(headers, "relevance summary", "summary", "relevance text")
    id_col       = _find_col(headers, " id")  # "id" alone is too broad — prefer "si id", " id"
    if id_col is None:
        id_col = _find_col(headers, "id")

    def cell(row, col):
        if col is not None and col < len(row) and row[col] is not None:
            return str(row[col]).strip()
        return ""

    total = high = medium = low = 0
    consistency_errors = 0
    lob_counter: Counter = Counter()
    category_counter: Counter = Counter()
    top_items: list = []

    for row in data_rows:
        total += 1
        title    = cell(row, title_col) or f"Item {total}"
        category = cell(row, category_col)
        relevance= cell(row, relevance_col)
        lob      = cell(row, lob_col)
        summary  = cell(row, summary_col)
        cons     = cell(row, status_col).lower()
        si_id    = cell(row, id_col)

        rank = _priority(category, relevance)
        if rank == 0:
            high += 1
        elif rank == 1:
            medium += 1
        else:
            low += 1

        if lob:
            lob_counter[lob] += 1
        if category:
            # Shorten for display
            cat_short = re.sub(r"\s*\(.*?\)", "", category).strip()
            category_counter[cat_short] += 1

        if any(t in cons for t in ("error", "fail", "incorrect", "inconsistent")):
            consistency_errors += 1

        label = "High" if rank == 0 else ("Medium" if rank == 1 else "Check")
        if len(top_items) < 10 or rank == 0:
            top_items.append({
                "title": title[:70],
                "category": category,
                "relevance": relevance,
                "lob": lob,
                "summary": summary[:150] if summary else "",
                "id": si_id,
                "priority": label,
                "rank": rank,
            })

    if total == 0:
        raise ValueError("No simplification items found in the export")

    top_items.sort(key=lambda x: x["rank"])

    # Derive intake fields (FUSE_ORDER puts SI lowest — overridden by RC/ATC if also uploaded).
    form: dict = {}

    mods = _lob_to_modules(lob_counter)
    if mods:
        form["modules_implemented"] = mods

    # process_reengineering_appetite: unavailable items signal forced redesign
    if high >= 5:
        form["process_reengineering_appetite"] = "redesign_to_standard"
    elif high >= 2:
        form["process_reengineering_appetite"] = "selective"

    facts = [
        f"{total} relevant Simplification Items  ·  {high} high  ·  {medium} medium  ·  {low} check",
    ]
    if lob_counter:
        top_lobs = ", ".join(f"{lob} ({n})" for lob, n in lob_counter.most_common(4))
        facts.append(f"By LoB: {top_lobs}")
    if category_counter:
        cats = ", ".join(f"{cat} ({n})" for cat, n in category_counter.most_common(3))
        facts.append(f"Categories: {cats}")
    if consistency_errors:
        facts.append(f"{consistency_errors} consistency error(s) — fix before conversion freeze")

    advisory: Optional[str] = None
    mandatory = [x for x in top_items if x["rank"] == 0]
    if mandatory:
        names = "; ".join(x["title"] for x in mandatory[:3])
        advisory = (
            f"Simplification Item Check flags {len(mandatory)} item(s) as 'Functionality "
            f"Unavailable' — these require mandatory action before conversion: {names}."
        )

    insights = {
        "kind": "simplification",
        "total": total,
        "high": high,
        "medium": medium,
        "low": low,
        "errors": consistency_errors,
        "lob_breakdown": lob_counter.most_common(),
        "category_breakdown": category_counter.most_common(),
        "items": top_items,
    }

    return {
        "form": form,
        "summary": {
            "facts": facts,
            "review": [
                "Mandatory 'Functionality Unavailable' items — plan before conversion",
                "Consistency errors — fix before conversion freeze",
                "Business inputs: driver, go-live, budget, risk, sovereignty, basis capability",
            ],
            "advisory": advisory,
        },
        "insights": insights,
    }
