"""Parse an SAP Readiness Check export (.zip of .xlsx + sizing txt, OR the
narrative .docx) into a pre-filled STAR intake plus an extraction summary.

Notes from real exports:
- The SAP-generated .xlsx set a wrong <dimension>, so openpyxl read_only mode
  under-reads. We load non-streaming (read_only=False) which reads all rows.
- Interface objects live in InterfaceImpactAnalysis{RFC,WEB,OData,BWE,IDOC}.xlsx,
  one object per data row.
- The .docx narrative report carries the same numbers in section tables; we match
  tables by header keywords (not index) so different export vintages still work.
"""
import io
import re
import zipfile
from typing import Optional

from openpyxl import load_workbook

_7Z_MAGIC = b"7z\xbc\xaf\x27\x1c"


def _extract_archive(data: bytes) -> dict:
    """Return {member_name: bytes} for any ZIP or 7z archive."""
    if data[:6] == _7Z_MAGIC:
        import py7zr, tempfile, os
        result: dict = {}
        try:
            with tempfile.TemporaryDirectory() as tmp:
                with py7zr.SevenZipFile(io.BytesIO(data), "r") as arc:
                    arc.extractall(path=tmp)
                for root, _, files in os.walk(tmp):
                    for fname in files:
                        full = os.path.join(root, fname)
                        rel = os.path.relpath(full, tmp).replace("\\", "/")
                        with open(full, "rb") as fh:
                            result[rel] = fh.read()
        except py7zr.exceptions.PasswordRequired:
            raise ValueError(
                "The .7z archive is password-protected. "
                "Please extract it with 7-Zip first, then upload the individual files."
            )
        return result
    z = zipfile.ZipFile(io.BytesIO(data))
    return {name: z.read(name) for name in z.namelist()
            if not name.split("/")[-1].startswith("~$")}

# upper bound (GiB) -> STAR band
_DB_BANDS = [
    (500, "lt_500gb"), (1000, "500gb_1tb"), (3000, "1_3tb"), (5000, "3_5tb"),
    (10000, "5_10tb"), (20000, "10_20tb"), (40000, "20_40tb"), (10 ** 12, "gt_40tb"),
]


def _band_from_gb(gb: float) -> str:
    for lim, name in _DB_BANDS:
        if gb < lim:
            return name
    return "gt_40tb"


def _rows(xlsx_bytes: bytes) -> list:
    try:
        wb = load_workbook(io.BytesIO(xlsx_bytes), read_only=False, data_only=True)
        ws = wb.worksheets[0]
        return [r for r in ws.iter_rows(values_only=True)
                if any(c is not None and str(c).strip() for c in r)]
    except Exception:
        return []


def _interface_band(total: int) -> str:
    if total < 20:
        return "lt_20"
    if total < 50:
        return "20_50"
    if total < 100:
        return "50_100"
    if total <= 200:
        return "100_200"
    return "gt_200"


def parse_zip(data: bytes) -> dict:
    """Return {'form': {...partial intake...}, 'summary': {...}}."""
    members = _extract_archive(data)
    names = list(members.keys())

    def find(substr: str) -> Optional[str]:
        for n in names:
            if substr.lower() in n.lower():
                return n
        return None

    form: dict = {}
    extracted: dict = {}   # exact numbers (where known) — UI shows these instead of bands
    facts: list = []
    review = ["SAP vs non-SAP interface split", "Middleware (PI/PO vs point-to-point)",
              "Business modules in scope",
              "Business inputs: driver, go-live, budget, risk, sovereignty, basis capability"]
    advisory: Optional[str] = None

    # --- system identity ---
    tp = find("TechnicalProperties")
    if tp:
        props = {}
        for r in _rows(members[tp]):
            if len(r) >= 2 and r[0]:
                props[str(r[0]).strip()] = (str(r[1]).strip() if r[1] is not None else "")
        inst = props.get("Installed Product Version", "")
        dbt = props.get("Database System Type", "")
        uni = props.get("Unicode", "")
        tgt = props.get("Target Product Version", "")
        sysid = (props.get("Analyzed System", "") or "SYS").split(" ")[0]
        is_hana = dbt.upper() in ("HDB", "HANA") or "hana" in dbt.lower()
        form["system_id"] = sysid
        form["unicode_status"] = "unicode" if uni.lower().startswith("y") else "non_unicode"
        if "S/4HANA" in inst:
            form["product_release"] = "s4hana"
            # Extract the release year (1503/1511/1610/1709/1809/1909/2020/…/2025) from inst.
            m_ver = re.search(r"S/4HANA[^\d]*(\d{4})", inst)
            if m_ver:
                form["s4_version"] = m_ver.group(1)
        elif is_hana:
            form["product_release"] = "soh"
        elif "6.0" in inst or "ERP 6" in inst:
            form["product_release"] = "ecc6_ehp"
        else:
            form["product_release"] = "ecc6"
        soh_tag = " already on HANA (Suite on HANA)" if is_hana and "S/4HANA" not in inst else ""
        facts.append(f"System {sysid} — {inst}{soh_tag}")
        facts.append(f"Unicode: {uni or 'n/a'} · Target: {tgt or 'n/a'}")

    # --- HANA sizing -> DB band ---
    sz = find("hana_sizing")
    if sz:
        txt = members[sz].decode("latin-1", "ignore")
        gb = None
        m = re.search(r"Data used size on disk[^\d]*(\d+)", txt)
        if m:
            gb = float(m.group(1))
        else:
            m = re.search(r"Net data volume size on disk\s+([\d.,]+)", txt)
            if m:
                gb = float(m.group(1).replace(".", "").replace(",", "."))
        if gb:
            form["db_size_band"] = _band_from_gb(gb)
            extracted["db_size_gb"] = int(gb)
            facts.append(f"HANA data size ≈ {int(gb)} GB")

    # --- custom code (delegated to the rich RC_CC parser for Quick Fix + topic insights) ---
    cc_insights = None
    cc = find("CustomCode")
    if cc:
        try:
            from app.engine.atc import _rows as _atc_rows, _parse_rc_cc, _is_rc_cc_format  # noqa: WPS433
            cc_rows = _atc_rows(members[cc])
            hdr_idx = 0
            cc_headers: list = []
            for i, row in enumerate(cc_rows[:8]):
                vals = [str(c).strip() for c in row if c is not None]
                if len(vals) >= 3 and not any(v.lstrip("-").replace(".", "").isdigit() for v in vals[:4]):
                    cc_headers = [str(c).lower().strip() if c else "" for c in row]
                    hdr_idx = i
                    break
            if _is_rc_cc_format(cc_headers):
                rc_cc = _parse_rc_cc(cc_rows[hdr_idx + 1:], cc_headers)
                form.update(rc_cc["form"])
                extracted.update(rc_cc.get("extracted") or {})
                facts.extend(rc_cc["summary"]["facts"])
                if rc_cc["summary"].get("advisory"):
                    advisory = rc_cc["summary"]["advisory"]
                cc_insights = rc_cc["insights"]
        except Exception:
            cc_insights = None

    if cc and cc_insights is None:
        rows = _rows(members[cc])
        findings = max(len(rows) - 1, 0)
        in_scope_total = 0
        try:
            hdr = [str(h).replace("\n", " ") for h in rows[0]]
            idx = next(i for i, h in enumerate(hdr) if "In Scope" in h)
            for r in rows[1:]:
                v = r[idx]
                try:
                    in_scope_total += int(float(str(v).replace(",", "").strip()))
                except (TypeError, ValueError):
                    pass
        except Exception:
            pass
        if in_scope_total >= 10000:
            form["custom_objects_band"] = "gt_10k"
        elif in_scope_total >= 2000:
            form["custom_objects_band"] = "2k_10k"
        elif in_scope_total >= 500:
            form["custom_objects_band"] = "500_2k"
        else:
            form["custom_objects_band"] = "lt_500"
        form["overall_customization"] = "High" if findings >= 40 else ("Med" if findings >= 12 else "Low")
        form["modifications_to_standard"] = "true"
        extracted["custom_objects"] = in_scope_total or findings
        facts.append(f"Custom code: {findings} simplification findings, ~{in_scope_total} impacted objects in scope")

    # --- interfaces ---
    total_if = 0
    if_breakdown: list[list] = []   # [[type, count], ...]
    for tech, key in [("RFC", "RFC"), ("WEB", "Web"), ("OData", "OData"), ("BWE", "BWE"), ("IDOC", "IDoc")]:
        n = find(f"InterfaceImpactAnalysis{tech}")
        if n:
            c = max(len(_rows(members[n])) - 1, 0)
            total_if += c
            if c:
                if_breakdown.append([key, c])
    if total_if:
        form["interface_count_band"] = _interface_band(total_if)
        form["interface_complexity"] = "very_complex" if total_if > 2000 else ("complex" if total_if > 500 else "medium")
        extracted["interface_count"] = total_if
        bd_str = " · ".join(f"{c} {k}" for k, c in if_breakdown)
        facts.append(f"Interfaces: {total_if:,} total ({bd_str})")

        # --- middleware inference from interface mix ---
        bd = {k: c for k, c in if_breakdown}
        web_n   = bd.get("Web",   0)
        odata_n = bd.get("OData", 0)
        rfc_n   = bd.get("RFC",   0)
        idoc_n  = bd.get("IDoc",  0)
        web_share   = web_n   / total_if
        odata_share = odata_n / total_if
        rfc_idoc_share = (rfc_n + idoc_n) / total_if
        if odata_share > 0.4 and web_share > 0.25:
            mw, mw_label = "mixed", "Mixed (modern OData + SOAP via PI/PO)"
        elif odata_share > 0.5:
            mw, mw_label = "sap_integration_suite", "SAP Integration Suite / CPI (OData-dominant)"
        elif web_share > 0.4:
            mw, mw_label = "sap_pi_po", "SAP PI / PO (SOAP-dominant via SOAMANAGER)"
        elif rfc_idoc_share > 0.7:
            mw, mw_label = "point_to_point", "Point-to-point (RFC/IDoc-dominant)"
        else:
            mw, mw_label = "mixed", "Mixed"
        form["middleware"] = mw
        facts.append(f"Middleware inferred: {mw_label}")

    # --- simplification items ---
    si_count = 0
    si = find("RelevantSimplificationItems")
    if si:
        si_count = max(len(_rows(members[si])) - 1, 0)
        if si_count:
            facts.append(f"Simplification items: {si_count} relevant")

    # --- add-on compatibility (conversion blocker) ---
    incompat = 0
    ad = find("AddonCompat")
    if ad:
        rows = _rows(members[ad])
        try:
            hdr = [str(h) for h in rows[0]]
            si2 = hdr.index("Status")
            incompat = sum(1 for r in rows[1:] if r[si2] and str(r[si2]).strip().lower() == "incompatible")
        except Exception:
            pass
        if incompat:
            advisory = (f"SAP Readiness Check flagged {incompat} incompatible add-on(s) — these block the "
                        f"SUM conversion and must be remediated or uninstalled as a prerequisite.")
            facts.append(f"Add-ons: {incompat} incompatible — conversion prerequisite")

    # --- Fiori recommended apps ---
    fiori_total = 0
    fiori_by_area: list[list] = []
    fa = find("RecommendedFioriApps")
    if fa:
        from collections import Counter as _Counter
        f_rows = _rows(members[fa])
        if len(f_rows) > 1:
            hdr_f = [str(h).lower() for h in f_rows[0]]
            area_col = next((i for i, h in enumerate(hdr_f) if "application area" in h or "area" in h), None)
            fiori_total = len(f_rows) - 1
            if area_col is not None:
                area_cnt = _Counter(str(r[area_col]).strip() for r in f_rows[1:] if r[area_col])
                fiori_by_area = [[a, n] for a, n in area_cnt.most_common(8) if a and a.lower() != "unknown"]
            if fiori_total >= 2000:
                form["fiori_apps_in_scope"] = "gt_2000"
            elif fiori_total >= 1000:
                form["fiori_apps_in_scope"] = "1000_2000"
            elif fiori_total >= 500:
                form["fiori_apps_in_scope"] = "500_1000"
            else:
                form["fiori_apps_in_scope"] = "lt_500"
            extracted["fiori_apps"] = fiori_total
            facts.append(f"Fiori: {fiori_total:,} recommended apps across {len(fiori_by_area)} functional areas")

    # --- Fiori activation level (from AppAvailability current UI tech mix) ---
    aa = find("AppAvailability")
    if aa:
        aa_rows = _rows(members[aa])
        if len(aa_rows) > 1:
            hdr_aa = [str(h) for h in aa_rows[0]]
            try:
                ui_i = hdr_aa.index("UI Technology Type")
                analyzed = [str(r[ui_i] or "").strip().lower() for r in aa_rows[1:] if r[ui_i]]
                total_aa = len(analyzed)
                fiori_count = sum(1 for u in analyzed if "fiori" in u)
                if total_aa:
                    pct = round(fiori_count / total_aa * 100)
                    if pct >= 75:
                        form["fiori_activation_level"] = "established"
                    elif pct >= 25:
                        form["fiori_activation_level"] = "expanding"
                    elif pct >= 1:
                        form["fiori_activation_level"] = "initial"
                    else:
                        form["fiori_activation_level"] = "not_started"
                    extracted["fiori_activation_pct"] = pct
                    facts.append(f"Fiori activation: {pct}% of {total_aa} in-use apps currently on Fiori UI")
            except (ValueError, IndexError):
                pass

    insights = {
        "kind": "readiness_zip",
        "interfaces": if_breakdown,
        "interfaceTotal": total_if,
        "fioriTotal": fiori_total,
        "fioriByArea": fiori_by_area,
        "siCount": si_count,
        "addonsIncompat": incompat,
    }
    if cc_insights:
        # Lift the rich Custom-Code details (Quick Fix, byCategory, byStatus, byComponent, totals)
        # into the RC insights so the frontend can render the same Quick Fix card it shows for ATC.
        insights["customCode"] = cc_insights
    return {"form": form, "summary": {"facts": facts, "review": review, "advisory": advisory},
            "insights": insights, "extracted": extracted}


# ---------- DOCX narrative parser ----------

def _docx_tables(data: bytes) -> list[list[list[str]]]:
    """Return every table as a list of rows; each row is a list of cell strings."""
    from docx import Document  # lazy import; only needed for docx uploads
    doc = Document(io.BytesIO(data))
    out: list[list[list[str]]] = []
    for tbl in doc.tables:
        rows = []
        for row in tbl.rows:
            rows.append([c.text.strip() for c in row.cells])
        out.append(rows)
    return out


def _find_table(tables: list[list[list[str]]], *needles: str) -> Optional[list[list[str]]]:
    """Return the first table whose header row contains every needle (case-insensitive)."""
    needles_l = [n.lower() for n in needles]
    for t in tables:
        if not t:
            continue
        header_blob = " | ".join(t[0]).lower()
        if all(n in header_blob for n in needles_l):
            return t
    return None


def _to_int(s: str) -> Optional[int]:
    if s is None:
        return None
    m = re.search(r"-?[\d,]+", s)
    if not m:
        return None
    try:
        return int(m.group(0).replace(",", ""))
    except ValueError:
        return None


def parse_docx(data: bytes) -> dict:
    """Extract intake fields from the SAP Readiness Check narrative .docx."""
    tables = _docx_tables(data)
    form: dict = {}
    facts: list[str] = []
    review = ["SAP vs non-SAP interface split", "Middleware (PI/PO vs point-to-point)",
              "Business modules in scope",
              "Business inputs: driver, go-live, budget, risk, sovereignty, basis capability"]
    advisory: Optional[str] = None

    # --- system identity ---
    sid_tbl = _find_table(tables, "system id", "date of analysis")
    if sid_tbl and len(sid_tbl) >= 2 and sid_tbl[1]:
        form["system_id"] = sid_tbl[1][0]
    prod_tbl = _find_table(tables, "system id", "installed product version")
    target_tbl = _find_table(tables, "product", "product version")
    installed = prod_tbl[1][1] if (prod_tbl and len(prod_tbl) >= 2) else ""
    target = ""
    if target_tbl and len(target_tbl) >= 2 and len(target_tbl[1]) >= 2:
        target = target_tbl[1][1]
    inst_l = installed.lower()
    if "s/4hana" in inst_l or "s4hana" in inst_l:
        form["product_release"] = "s4hana"
        # S/4HANA is always Unicode
        form["unicode_status"] = "unicode"
        form["stack_type"] = "single_stack"
        m = re.search(r"(1\d{3}|2\d{3})", installed)
        if m:
            form["s4_version"] = m.group(1)
    db_tbl = _find_table(tables, "system id", "database system")
    if db_tbl and len(db_tbl) >= 2 and len(db_tbl[1]) >= 2:
        dbsys = db_tbl[1][1]
        if dbsys.upper() in ("HDB", "HANA") or "hana" in dbsys.lower():
            if form.get("product_release") != "s4hana":
                form["product_release"] = "soh"
                form["unicode_status"] = "unicode"
    if installed:
        facts.append(f"System {form.get('system_id', '')} — {installed}"
                     + (f" → target {target}" if target else ""))

    # --- custom code findings ---
    cc_tbl = _find_table(tables, "item", "findings", "unresolved")
    if cc_tbl:
        total_unresolved = 0
        in_scope_manual = 0
        for row in cc_tbl[1:]:
            if not row:
                continue
            label = row[0].lower()
            v = _to_int(row[1]) if len(row) > 1 else None
            if v is None:
                continue
            if label.startswith("total findings"):
                total_unresolved = v
            elif "in scope" in label and "manual" in label.lower():
                in_scope_manual += v
        if in_scope_manual >= 10000:
            form["custom_objects_band"] = "gt_10k"
        elif in_scope_manual >= 2000:
            form["custom_objects_band"] = "2k_10k"
        elif in_scope_manual >= 500:
            form["custom_objects_band"] = "500_2k"
        elif in_scope_manual > 0:
            form["custom_objects_band"] = "lt_500"
        if in_scope_manual >= 2000:
            form["overall_customization"] = "High"
        elif in_scope_manual >= 500:
            form["overall_customization"] = "Med"
        elif in_scope_manual > 0:
            form["overall_customization"] = "Low"
        if total_unresolved or in_scope_manual:
            facts.append(f"Custom code: {total_unresolved} unresolved findings "
                         f"(~{in_scope_manual} in-scope manual)")

    # --- interfaces (Total Number | Impacted | Mediated via Middleware) ---
    if_tbl = _find_table(tables, "total number", "impacted", "mediated")
    if if_tbl and len(if_tbl) >= 2:
        total_if = _to_int(if_tbl[1][0])
        if total_if:
            form["interface_count_band"] = _interface_band(total_if)
            form["interface_complexity"] = "very_complex" if total_if > 2000 else (
                "complex" if total_if > 500 else "medium")
            facts.append(f"Interfaces: ≈ {total_if} objects analyzed")

    # --- simplification items (relevant count) ---
    si_tbl = _find_table(tables, "status", "number of simplification items")
    if si_tbl:
        relevant = 0
        for row in si_tbl[1:]:
            if len(row) < 2:
                continue
            label = row[0].lower()
            v = _to_int(row[1])
            if v is None:
                continue
            if "is relevant" in label or label.startswith("check performed") and "is relevant" in label:
                relevant += v
        if relevant:
            facts.append(f"Simplification items: {relevant} relevant")

    # --- add-on compatibility (status -> count) ---
    ad_tbl = _find_table(tables, "status", "number of add-ons")
    if ad_tbl:
        incompat = 0
        for row in ad_tbl[1:]:
            if len(row) < 2:
                continue
            label = row[0].lower()
            v = _to_int(row[1]) or 0
            if "incompatible" in label:
                incompat += v
        if incompat:
            advisory = (f"SAP Readiness Check flagged {incompat} incompatible add-on(s) — these block the "
                        f"SUM conversion and must be remediated or uninstalled as a prerequisite.")
            facts.append(f"Add-ons: {incompat} incompatible — conversion prerequisite")

    # kind "readiness_docx" keeps the null-check in the frontend from crashing —
    # the zip-based "readiness" panel expects interface/customCode/addons arrays
    # which the docx doesn't carry.
    insights = {"kind": "readiness_docx", "facts": facts}
    return {"form": form, "summary": {"facts": facts, "review": review, "advisory": advisory}, "insights": insights}


def parse(data: bytes, filename: str) -> dict:
    """Dispatch on extension: .docx → narrative parser, .zip/.7z → archive parser."""
    name = (filename or "").lower()
    if name.endswith(".docx"):
        return parse_docx(data)
    return parse_zip(data)  # handles both .zip and .7z via _extract_archive
